import io
import uuid
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from flask import (Blueprint, jsonify, redirect, render_template,
                   request, send_file, session, url_for)
import db.queries as q

teacher_bp = Blueprint("teacher", __name__)


# ── HTML login (standalone Flask use) ────────────────────────
@teacher_bp.route("/teacher-login", methods=["GET", "POST"])
def teacher_login():
    if request.method == "POST":
        t = q.get_teacher_by_email(request.form.get("email", ""))
        if t and t["password"] == request.form.get("password", ""):
            session["teacher"]    = t["email"]
            session["teacher_id"] = t["id"]
            return redirect(url_for("teacher.teacher_dashboard"))
        return render_template("teacher_login.html", error="Invalid Credentials")
    return render_template("teacher_login.html")


# ── JSON login for React ──────────────────────────────────────
@teacher_bp.route("/teacher-login-json", methods=["POST", "OPTIONS"])
def teacher_login_json():
    if request.method == "OPTIONS": return "", 204
    data = request.json or {}
    t    = q.get_teacher_by_email(data.get("email", ""))
    if t and t["password"] == data.get("password", ""):
        session["teacher"]    = t["email"]
        session["teacher_id"] = t["id"]
        name = t["email"].split("@")[0].replace(".", " ").title()
        return jsonify({
            "user" : {"id": t["id"], "name": name, "email": t["email"], "role": "teacher"},
            "token": "flask-session"
        })
    return jsonify({"error": "Invalid credentials"}), 401


@teacher_bp.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("teacher.teacher_login"))


@teacher_bp.route("/teacher-dashboard")
def teacher_dashboard():
    if "teacher" not in session: return redirect(url_for("teacher.teacher_login"))
    return render_template("teacher_dashboard.html")


# ── Session management ────────────────────────────────────────
@teacher_bp.route("/start-session", methods=["POST", "OPTIONS"])
def start_session():
    if request.method == "OPTIONS": return "", 204
    data     = request.json or {}
    date_str = data.get("date")
    sess_num = int(data.get("session_number", 1))
    subject  = data.get("subject",  "General")
    semester = data.get("semester", "Semester 1")

    if not date_str: return jsonify({"error": "date is required"}), 400

    # Initialise all students as Absent for this exact subject+semester session
    q.init_attendance_for_session(date_str, sess_num, subject, semester)

    # Create session record in DB (lazy rotation replaces background thread)
    session_uuid = str(uuid.uuid4())
    token        = str(uuid.uuid4())
    q.create_db_session(session_uuid, token, date_str, sess_num, subject, semester)

    return jsonify({"status": "started", "date": date_str,
                    "session_number": sess_num, "subject": subject, "semester": semester})


@teacher_bp.route("/current-token")
def get_token():
    sess = q.get_active_db_session()
    if not sess: return jsonify({"session_id": None, "token": None})
    # Lazy rotation — rotate token in-request instead of background thread
    sess = q.rotate_token_if_needed(sess["session_uuid"])
    if not sess: return jsonify({"session_id": None, "token": None})
    return jsonify({"session_id": sess["session_uuid"], "token": sess["current_token"]})


@teacher_bp.route("/stop-session", methods=["POST", "OPTIONS"])
def stop_session():
    if request.method == "OPTIONS": return "", 204
    q.stop_db_session()
    return jsonify({"status": "stopped"})


# ── Live count ────────────────────────────────────────────────
@teacher_bp.route("/attendance-count")
def attendance_count():
    sess = q.get_active_db_session()
    if not sess: return jsonify({"marked": 0, "total": 0, "session": ""})
    m, t = q.count_marked(
        str(sess["session_date"]), sess["session_number"],
        sess["subject"], sess["semester"]
    )
    return jsonify({
        "marked" : m, "total": t,
        "session": f"{sess['semester']} · {sess['subject']} · {sess['session_date']} S{sess['session_number']}"
    })


# ── Attendance table ──────────────────────────────────────────
@teacher_bp.route("/get-attendance-table")
def get_attendance_table():
    if "teacher" not in session: return jsonify({"error": "Unauthorized"}), 401
    subject  = request.args.get("subject")
    semester = request.args.get("semester")
    return jsonify(q.get_attendance_table(subject, semester))


# ── Delete ────────────────────────────────────────────────────
@teacher_bp.route("/delete-attendance", methods=["DELETE", "OPTIONS"])
def delete_attendance():
    if request.method == "OPTIONS": return "", 204
    if "teacher" not in session: return jsonify({"error": "Unauthorized"}), 401
    data = request.json or {}
    sid  = data.get("student_id")
    if not sid: return jsonify({"error": "student_id required"}), 400
    d, sn = data.get("date"), data.get("session_number")
    if d and sn is not None: q.delete_specific_attendance(int(sid), d, int(sn))
    else:                    q.delete_all_attendance_for_student(int(sid))
    return jsonify({"status": "deleted"})


@teacher_bp.route("/delete-attendance-session", methods=["DELETE", "OPTIONS"])
def delete_attendance_session():
    if request.method == "OPTIONS": return "", 204
    if "teacher" not in session: return jsonify({"error": "Unauthorized"}), 401
    data = request.json or {}
    q.delete_entire_date_session(data.get("date"), int(data.get("session_number", 1)))
    return jsonify({"status": "deleted"})


# ── Export Excel ──────────────────────────────────────────────
@teacher_bp.route("/export-excel")
def export_excel():
    if "teacher" not in session: return redirect(url_for("teacher.teacher_login"))
    subject  = request.args.get("subject")
    semester = request.args.get("semester")
    data     = q.get_attendance_table(subject, semester)

    wb  = openpyxl.Workbook(); ws = wb.active; ws.title = "Attendance"
    hf  = PatternFill("solid", fgColor="1E3C72"); hfnt = Font(color="FFFFFF", bold=True)
    pf  = PatternFill("solid", fgColor="C6EFCE"); af   = PatternFill("solid", fgColor="FFC7CE")
    ctr = Alignment(horizontal="center", vertical="center")

    headers = ["SL No.", "Reg No.", "Name"] + [c.replace("_S", " S") for c in data["columns"]]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h); cell.fill=hf; cell.font=hfnt; cell.alignment=ctr

    for rd in data["rows"]:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=rd["sl"]).alignment  = ctr
        ws.cell(row=r, column=2, value=rd["reg_no"]).font   = Font(bold=True)
        ws.cell(row=r, column=3, value=rd["name"])
        for ci, ck in enumerate(data["columns"], 4):
            v    = rd.get(ck, "–"); cell = ws.cell(row=r, column=ci, value=v); cell.alignment = ctr
            if v == "P": cell.fill = pf; cell.font = Font(bold=True, color="375623")
            elif v == "A": cell.fill = af; cell.font = Font(bold=True, color="9C0006")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 6, 30)
    ws.freeze_panes = "A2"

    out = io.BytesIO(); wb.save(out); out.seek(0)
    fname = f"attendance_{(subject or 'all').replace(' ','_')}.xlsx"
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname)